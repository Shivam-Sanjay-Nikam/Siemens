#!/usr/bin/env python3
"""
Script to import sample data into the Order Management System database.
This will populate the database with sample employees and items for testing.
"""

import sys
import os
import csv
import openpyxl
from db import Database

def import_employees_from_csv(db, file_path):
    """Import employees from CSV file."""
    print(f"Importing employees from {file_path}...")
    imported_count = 0
    
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            header = next(reader)  # Skip header row
            
            for row in reader:
                if len(row) >= 2:
                    emp_id = row[0].strip()
                    emp_name = row[1].strip()
                    
                    if emp_id and emp_name:
                        try:
                            db.add_employee(emp_id, emp_name)
                            imported_count += 1
                            print(f"  ✓ Added employee: {emp_id} - {emp_name}")
                        except Exception as e:
                            print(f"  ✗ Failed to add employee {emp_id}: {e}")
                            
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        
    print(f"Imported {imported_count} employees from CSV")
    return imported_count

def import_employees_from_excel(db, file_path):
    """Import employees from Excel file."""
    print(f"Importing employees from {file_path}...")
    imported_count = 0
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            if row[0] and row[1]:
                emp_id = str(row[0]).strip()
                emp_name = str(row[1]).strip()
                
                try:
                    db.add_employee(emp_id, emp_name)
                    imported_count += 1
                    print(f"  ✓ Added employee: {emp_id} - {emp_name}")
                except Exception as e:
                    print(f"  ✗ Failed to add employee {emp_id}: {e}")
                    
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        
    print(f"Imported {imported_count} employees from Excel")
    return imported_count

def import_items_from_csv(db, file_path):
    """Import items from CSV file."""
    print(f"Importing items from {file_path}...")
    imported_count = 0
    
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            header = next(reader)  # Skip header row
            
            for row in reader:
                if len(row) >= 2:
                    name = row[0].strip()
                    price_str = row[1].strip()
                    
                    try:
                        price = float(price_str)
                        if name and price > 0:
                            db.add_item(name, price)
                            imported_count += 1
                            print(f"  ✓ Added item: {name} - ₹{price}")
                    except ValueError:
                        print(f"  ✗ Invalid price for {name}: {price_str}")
                        
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        
    print(f"Imported {imported_count} items from CSV")
    return imported_count

def import_items_from_excel(db, file_path):
    """Import items from Excel file."""
    print(f"Importing items from {file_path}...")
    imported_count = 0
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            if row[0] and row[1]:
                name = str(row[0]).strip()
                try:
                    price = float(row[1])
                    if name and price > 0:
                        db.add_item(name, price)
                        imported_count += 1
                        print(f"  ✓ Added item: {name} - ₹{price}")
                except ValueError:
                    print(f"  ✗ Invalid price for {name}: {row[1]}")
                    
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        
    print(f"Imported {imported_count} items from Excel")
    return imported_count

def main():
    """Main function to import all sample data."""
    print("=" * 60)
    print("ORDER MANAGEMENT SYSTEM - SAMPLE DATA IMPORT")
    print("=" * 60)
    
    # Initialize database
    db = Database()
    
    # Import employees
    print("\n📋 IMPORTING EMPLOYEES")
    print("-" * 30)
    
    # Try CSV first, then Excel
    if os.path.exists("sample_employees.csv"):
        import_employees_from_csv(db, "sample_employees.csv")
    elif os.path.exists("sample_employees.xlsx"):
        import_employees_from_excel(db, "sample_employees.xlsx")
    else:
        print("❌ No sample employee files found!")
    
    # Import items
    print("\n🍽️ IMPORTING ITEMS")
    print("-" * 30)
    
    # Try CSV first, then Excel
    if os.path.exists("sample_items.csv"):
        import_items_from_csv(db, "sample_items.csv")
    elif os.path.exists("sample_items.xlsx"):
        import_items_from_excel(db, "sample_items.xlsx")
    else:
        print("❌ No sample item files found!")
    
    # Set some items as today's menu
    print("\n📅 SETTING TODAY'S MENU")
    print("-" * 30)
    
    try:
        items = db.get_items()
        if items:
            # Set first 10 items as today's menu
            today_items = items[:10]
            item_ids = [item[0] for item in today_items]
            db.set_today_menu(item_ids)
            print(f"✓ Set {len(item_ids)} items as today's menu:")
            for item_id, name, price in today_items:
                print(f"  • {name} - ₹{price}")
        else:
            print("❌ No items found to set as today's menu!")
    except Exception as e:
        print(f"❌ Error setting today's menu: {e}")
    
    # Show summary
    print("\n📊 IMPORT SUMMARY")
    print("-" * 30)
    
    try:
        employees = db.get_employees()
        items = db.get_items()
        today_menu = db.get_today_menu()
        
        print(f"✓ Total employees: {len(employees)}")
        print(f"✓ Total items: {len(items)}")
        print(f"✓ Today's menu items: {len(today_menu)}")
        
        if employees:
            print(f"\n👥 Sample employees:")
            for emp_id, name, due in employees[:5]:  # Show first 5
                print(f"  • {emp_id} - {name} (Due: ₹{due:.2f})")
            if len(employees) > 5:
                print(f"  ... and {len(employees) - 5} more")
                
        if today_menu:
            print(f"\n🍽️ Today's menu:")
            for item_id, name, price in today_menu[:5]:  # Show first 5
                print(f"  • {name} - ₹{price}")
            if len(today_menu) > 5:
                print(f"  ... and {len(today_menu) - 5} more")
                
    except Exception as e:
        print(f"❌ Error getting summary: {e}")
    
    print("\n" + "=" * 60)
    print("✅ SAMPLE DATA IMPORT COMPLETED!")
    print("You can now test the application with the imported data.")
    print("=" * 60)

if __name__ == "__main__":
    main()
